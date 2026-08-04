import io
import json
import os
import re
from datetime import datetime
from http.server import BaseHTTPRequestHandler
from urllib.parse import quote

from openpyxl import load_workbook


MASTER_FILE = os.path.join(os.getcwd(), "SiamTin_Master.xlsx")
MASTER_SHEET = "ONE (2)"


# Plan Load field -> Excel Master cell
CELL_MAP = {
    "FA NO.": "D6",
    "INVOICE NO.": "F6",

    "FROM": "D10",
    "TO": "H10",
    "DRIVER NAME": "D11",

    "CONT NO.": "D12",
    "BOOKING NO": "D13",
    "Liner": "D14",
    "SEAL NO.": "D15",

    "TRUCK NO.": "J13",
    "TRAILER NO.": "J14",
    "DRIVER PHONE": "I15",

    "PICKUP CONTACT": "C16",
    "PICKUP PLACE": "C17",
    "PICKUP DATE": "H17",

    "FACTORY DATE": "K17",
    "FACTORY TIME": "M17",

    "RETURN PLACE": "C18",
    "Return date": "H18",
    "RETURN TIME": "M18",

    "APPROVER": "K19",
    "APPROVE DATE": "K20",
}


DATE_FIELDS = {
    "PICKUP DATE",
    "FACTORY DATE",
    "Return date",
    "APPROVE DATE",
}

TIME_FIELDS = {
    "FACTORY TIME",
    "RETURN TIME",
}


def clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def parse_date(value):
    """
    Return a real Python date where possible.
    If the input cannot be parsed, keep the original text.
    """
    text = clean_text(value)
    if not text:
        return ""

    formats = (
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%Y-%m-%d %H:%M:%S",
    )

    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass

    return text


def parse_time(value):
    """
    Return a real Python time where possible.
    """
    text = clean_text(value)
    if not text:
        return ""

    match = re.search(r"(\d{1,2}):(\d{2})", text)
    if not match:
        return text

    hour = int(match.group(1))
    minute = int(match.group(2))

    if 0 <= hour <= 23 and 0 <= minute <= 59:
        return datetime(1899, 12, 30, hour, minute).time()

    return text


def safe_filename_part(value):
    text = clean_text(value)
    text = re.sub(r'[\\/:*?"<>|]+', "-", text)
    text = re.sub(r"\s+", "_", text)
    return text[:60]


def set_cell_value(ws, cell, field, value):
    if field in DATE_FIELDS:
        parsed = parse_date(value)
        ws[cell] = parsed

        if parsed and not isinstance(parsed, str):
            ws[cell].number_format = "dd/mm/yyyy"

        return

    if field in TIME_FIELDS:
        parsed = parse_time(value)
        ws[cell] = parsed

        if parsed and not isinstance(parsed, str):
            ws[cell].number_format = "hh:mm"

        return

    ws[cell] = clean_text(value)


def apply_transport_company(ws, data):
    """
    The company master has:
      row 8 = รถบรรทุก บริษัท
      row 9 = หัวลาก บริษัท

    Fill only the appropriate row when TRUCK TYPE is supplied.
    If TRUCK TYPE is blank, use the truck row as the default.
    """
    truck_type = clean_text(data.get("TRUCK TYPE")).upper()
    company = clean_text(data.get("Truck Com."))
    phone = clean_text(data.get("DRIVER PHONE"))

    is_trailer = (
        "TRAILER" in truck_type
        or "หัวลาก" in truck_type
    )

    if is_trailer:
        ws["D8"] = ""
        ws["H8"] = ""
        ws["D9"] = company
        ws["H9"] = phone
    else:
        ws["D8"] = company
        ws["H8"] = phone
        ws["D9"] = ""
        ws["H9"] = ""


def build_excel(data):
    if not os.path.exists(MASTER_FILE):
        raise FileNotFoundError(
            "ไม่พบไฟล์ SiamTin_Master.xlsx ใน GitHub/Vercel project"
        )

    # Load a fresh copy on every request.
    # The original Master file is never overwritten.
    wb = load_workbook(
        MASTER_FILE,
        data_only=False,
        keep_links=True,
    )

    if MASTER_SHEET not in wb.sheetnames:
        raise ValueError(
            f'ไม่พบชีต "{MASTER_SHEET}" ใน SiamTin_Master.xlsx'
        )

    ws = wb[MASTER_SHEET]

    for field, cell in CELL_MAP.items():
        set_cell_value(
            ws,
            cell,
            field,
            data.get(field, ""),
        )

    apply_transport_company(ws, data)

    # V2 intentionally DOES NOT add X / checkbox marks.
    # This keeps the original form clean while we first confirm
    # the generated workbook opens normally in Excel.

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()


class handler(BaseHTTPRequestHandler):

    def _cors(self):
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header(
            "Access-Control-Allow-Methods",
            "GET, POST, OPTIONS"
        )
        self.send_header(
            "Access-Control-Allow-Headers",
            "Content-Type"
        )

    def do_OPTIONS(self):
        self.send_response(204)
        self._cors()
        self.end_headers()

    def do_GET(self):
        body = json.dumps(
            {
                "ok": True,
                "service": "Siam Tin Excel Export",
                "engine": "Python + openpyxl",
                "master": "SiamTin_Master.xlsx",
                "sheet": MASTER_SHEET,
            },
            ensure_ascii=False,
        ).encode("utf-8")

        self.send_response(200)
        self._cors()
        self.send_header(
            "Content-Type",
            "application/json; charset=utf-8"
        )
        self.send_header(
            "Content-Length",
            str(len(body))
        )
        self.end_headers()
        self.wfile.write(body)

    def do_POST(self):
        try:
            content_length = int(
                self.headers.get("Content-Length", "0")
            )

            raw = self.rfile.read(content_length)

            data = (
                json.loads(raw.decode("utf-8"))
                if raw
                else {}
            )

            excel_bytes = build_excel(data)

            fa = safe_filename_part(
                data.get("FA NO.", "SiamTin")
            ) or "SiamTin"

            booking = safe_filename_part(
                data.get("BOOKING NO", "")
            )

            filename = (
                f"{fa}_{booking}.xlsx"
                if booking
                else f"{fa}.xlsx"
            )

            self.send_response(200)
            self._cors()
            self.send_header(
                "Content-Type",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            self.send_header(
                "Content-Disposition",
                "attachment; "
                f"filename*=UTF-8''{quote(filename)}"
            )
            self.send_header(
                "Content-Length",
                str(len(excel_bytes))
            )
            self.end_headers()
            self.wfile.write(excel_bytes)

        except Exception as error:
            message = json.dumps(
                {
                    "ok": False,
                    "error": str(error),
                },
                ensure_ascii=False,
            ).encode("utf-8")

            self.send_response(500)
            self._cors()
            self.send_header(
                "Content-Type",
                "application/json; charset=utf-8"
            )
            self.send_header(
                "Content-Length",
                str(len(message))
            )
            self.end_headers()
            self.wfile.write(message)
